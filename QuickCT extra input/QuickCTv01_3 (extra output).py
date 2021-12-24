# Cycle Tempo automisation program
#
# Author: Berend van Veldhuizen
# Version 0.1
# Start date: 2021-08-24

# import packages
from tkinter import *                                               # input box package
from tkinter import filedialog                                      # input file search package
import subprocess                                                   # package running other .exe
import os                                                           # package for deleting files
#import os, sys, win32com.shell.shell as shell                      # packages for running as Administrator
import pandas as pd                                                 # reading and writing to excel
import numpy as np                                                  # creating lists
from openpyxl import load_workbook, Workbook
import numpy as np
import getpass                                                      # package to retrieve current username
import time                                                         # package to pause time
from datetime import datetime                                       # package to retrieve current time


# import own functions
import functions

# define names of input and output file
#excel_input_name = "CT_input"
#excel_output_name = "CT_output"

# predefine variables (will be defined by user)
apparatus_def = 6
var_name_def = "VCELL"


# Define input window
window = Tk()
window.title("Cycle Tempo Automatisation")
window.geometry('1000x600')

def button_enter_CTpath():
    CTpath = filedialog.askopenfilename()
    entry1.insert(0,CTpath)
    print(CTpath)
    return None

def button_enter_EXCELpath():
    EXCELpath = filedialog.askopenfilename()
    entry4.insert(0,EXCELpath)
    print(EXCELpath)
    return None

# ----- startbutton function -----
def button_command_start():
    # Make sure no OUTFIL exists in folder, since CTm cannot overwrite
    OUTFILES = "OUTFIL1","OUTFIL2","OUTFIL3","OUTFIL4"
    
    print("Cleaning current folder:")
    for file in OUTFILES:
        if os.path.isfile(file):
            os.remove(file)
            print("   Existing ", file, " removed")
        else:
            print("   No ", file,"had to be removed from folder")
    
    print("Used input:")
    # Cycle tempo path
    CTpath = entry1.get()           # Get path
    if bool(CTpath):
        print("   Used entry CT path: ", CTpath)
    else:
    # Pre-define path so don't have to enter it every time
        CTpath = 'C:\\Program Files (x86)\\Cycle-Tempo\\WinTempo.exe'
        print("   Used predefined CT path: ", CTpath)
    # Excel path
    EXCELpath = entry4.get()        # Get path
    if bool(EXCELpath):
        print("   Used entry EXCEL path: ", EXCELpath)
    else:
    # Pre-define path so don't have to enter it every time
        #username = getpass.getuser()
        #EXCELpath = 'C:\\Users\\'+username+'\\surfdrive\\0. PHD (surfdrive)\\02 Research\\Cycle tempo\\2021-08-24 CT automisation\\CT_input.xlsx'
        EXCELpath = 'CT_input.xlsx'
        print("   Used predefined EXCEL path: ", EXCELpath)
    # Retrieve apparatus number        
    apparatus = entry2.get()
    if bool(apparatus):
        print("   Used apparatus number: ",apparatus)
    else:
        apparatus = str(apparatus_def)
        print("   Used predifined apparatus number: ", apparatus_def)
    # Retrieve variable name
    var_name = entry3.get()
    if bool(var_name):
        print("   Used variable: ",var_name, "\n")
    else:
        var_name = var_name_def
        print("   Used predifined variable:", var_name_def, "\n")
    # ----- Retrieve input choice and input parameters -----
    input_choice = int(input_choice_tk.get())
    #print(input_choice)
    #print(type(input_choice))
    if input_choice == 1:
        print("Excel is used as input")
        # Extract varying parameter values from excel s
        var_values = pd.read_excel(EXCELpath, sheet_name="Input")
        print(var_values, '\n')
        #print(type(var_values))
        var_values = var_values['Varying parameter'].tolist()    #convert DF to list to use in iteration        
    else:
        print("Given range is used as input")
        # Creat list from entered range values
        value_start = float(entry5.get())
        #print(value_start)
        #print(type(value_start))
        value_end = float(entry6.get())
        stepsize = float(entry7.get())
         
        var_values = np.arange(value_start, value_end, stepsize)
        print("With values:", var_values,"\n")
    # ----- Retrieve output choice -----
    output_choice = [0,0,0,0,0,0,0]
    output_choice[0] = additional_output1_var.get()
    output_choice[1] = additional_output2_var.get()
    output_choice[2] = additional_output3_var.get()
    output_choice[3] = additional_output4_var.get()
    output_choice[4] = additional_output5_var.get()
    output_choice[5] = additional_output6_var.get()
    output_choice[6] = additional_output7_var.get()
    print(output_choice)
        
    # ----- START ITERATION -----
    # Create empty dataframe for output
    output = pd.DataFrame()
    
    for var_value in var_values:
        print('Variable value of current iteration:',var_value)   
        # ----- Substitute variable value in INFILE1 -----
        var_value_span = functions.change_INFILE(apparatus,var_name,var_value)
      
        # ----- Run cycle tempo -----
        subprocess.Popen(CTpath)                     # run CT calculation
        time.sleep(0.5)                              # pause script otherwise the results are not generated yet
        # Extract results from OUTFIL4
        efficiencies, additional_output = functions.extract_OUTFILE(output_choice)
        # Remove pop-up window # to be added

        # Delete OUTFIL files to make space for next iteration
        for file in OUTFILES:
            os.remove(file)
    
        # ----- Process data -----
        # Append to [nx8] dataframe
        efficiencies_oneline = pd.DataFrame([y.values.ravel() for x , y in efficiencies.groupby(np.arange(len(efficiencies))//2)])      #transform to 1x8 df
        efficiencies_oneline.insert(0, var_name+"(apparatus "+ apparatus+")",var_value)
        efficiencies_oneline = efficiencies_oneline.rename(columns={0:'Gross energy efficiency',1:'Net energy efficiency',2:'Heat energy efficiency',3:'Total energy efficiency',4:'Gross exergy efficiency',5:'Net exergy efficiency',6:'Heat exergy efficiency',7:'Total exergy efficiency'})
        #print(efficiencies_oneline)
             
        output_oneline = efficiencies_oneline                  
        if output_choice[0] != '0':
            output_oneline = output_oneline.assign(SOFC_AC_power = additional_output[0])
            
        if output_choice[1] != '0':
            output_oneline = output_oneline.assign(SOFC_U_O = additional_output[1])   
            
        if output_choice[2] != '0':
            output_oneline = output_oneline.assign(SOFC_i = additional_output[2])        
            
        if output_choice[3] != '0':
            output_oneline = output_oneline.assign(Air_comp_P = additional_output[3])    
            
        if output_choice[4] != '0':            
            output_oneline = output_oneline.assign(Air_preh_Phi = additional_output[4])      
            
        if output_choice[5] != '0':            
            output_oneline = output_oneline.assign(Sink_ST_Phi = additional_output[5])    
            
        if output_choice[6] != '0':
            output_oneline = output_oneline.assign(Sink_HW_Phi = additional_output[6])    
            
        
        
        print(output_oneline)
        output = output.append(output_oneline)                                           # append row of iteration to output

        print(output)
        output_formatted = '=' + output.astype(str)                                   # add =  so excel automatticaly makes number of the strings
   
    # @@@@@@ END ITERATION @@@@@
    
    # ----- Write to excel -----
    # add column names
    
    # define current date and time for worksheet name
    now = datetime.now()
    dt_string = now.strftime("%y-%m-%d %H_%M")
    print("date and time =", dt_string)   
    # check whether CT_output exist and create new one if not
    if not os.path.isfile("CT_output.xlsx"):
        print("CT_output file did not exist, new one has been created and new worksheet has been added")
        wb = Workbook()
        wb.save(filename = 'CT_output.xlsx')
    else:
        print("CT_output file did exist, new worksheet has been added")
    book = load_workbook('CT_output.xlsx')
    writer = pd.ExcelWriter('CT_output.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    output_formatted.to_excel(writer, dt_string, index = False)
    writer.save()    
    
    return None

# clearbutton function
def button_command_clear():
    entry1.delete(0, END)
    entry2.delete(0, END)
    entry3.delete(0, END)
    entry4.delete(0, END)
    entry5.delete(0, END)
    entry6.delete(0, END)    
    entry7.delete(0, END)    
    return None



# ----- Text, entries and buttons (grid is used for placement) -----
# Cycle Tempo path entry
label1 = Label(window, text='Enter WinTempo.exe Path: ', fg='black', font=('Arial',12))
label1.grid(row=0,column=0,padx=5,pady=10, sticky=W)

entry1 = Entry(window, width = 50, font=('Arial',9))
entry1.grid(row=0, column=1)

CT_path_button = Button(window, text="...", command = button_enter_CTpath, width = 3)
CT_path_button.grid(row=0, column=2, padx=5, sticky=W)

# Cycle Tempo information
label2 = Label(window, text='Apparatus number: ', fg='black', font=('Arial',12))
label2.grid(row=1,column=0,padx=5,pady=10, sticky=W)

entry2 = Entry(window, width = 20, font=('Arial',12))
entry2.grid(row=1, column=1, sticky=W)

label3 = Label(window, text='Varying parameter name: ', fg='black', font=('Arial',12))
label3.grid(row=2,column=0,padx=5,pady=10, sticky=W)

entry3 = Entry(window, width = 20, font=('Arial',12))
entry3.grid(row=2, column=1, sticky=W)

def ExcelOption():
    entry4.configure(state="normal")
    entry4.update()    
    entry5.configure(state="disabled")
    entry5.update()
    entry6.configure(state="disabled")
    entry6.update()    
    entry7.configure(state="disabled")
    entry7.update()
def RangeOption(): 
    entry4.configure(state="disabled")
    entry4.update()    
    entry5.configure(state="normal")
    entry5.update()
    entry6.configure(state="normal")
    entry6.update()    
    entry7.configure(state="normal")
    entry7.update()

input_choice_tk = StringVar()
input_choice_tk.set("1")
option1 = Radiobutton(window, text = "Use excel as input", variable = input_choice_tk, value="1", command=ExcelOption, fg='black', font=('Arial',12))
option1.grid(row=3, column=0, sticky=W)

# Excel path entry
label4 = Label(window, text='Enter Excel Path: ', fg='black', font=('Arial',12))
label4.grid(row=4,column=0,padx=5,pady=10, sticky=E)

entry4 = Entry(window, width = 50, font=('Arial',9))
entry4.grid(row=4, column=1)

EXCEL_path_button = Button(window, text="...", command = button_enter_EXCELpath, width = 3)
EXCEL_path_button.grid(row=4, column=2, padx=5, sticky=W)

option2 = Radiobutton(window, text = "Give range as input:", variable = input_choice_tk, value="2", command=RangeOption, fg='black', font=('Arial',12))
option2.grid(row=5, column=0, sticky=W)

label5 = Label(window, text='Start value: ', fg='black', font=('Arial',12))
label5.grid(row=6,column=0,padx=5,pady=10, sticky=E)
entry5 = Entry(window, width = 10, font=('Arial',12))
entry5.grid(row=6, column=1, sticky=W)
label6 = Label(window, text='End value: ', fg='black', font=('Arial',12))
label6.grid(row=7,column=0,padx=5,pady=10, sticky=E)
entry6 = Entry(window, width = 10, font=('Arial',12))
entry6.grid(row=7, column=1, sticky=W)
label7 = Label(window, text='Stepsize: ', fg='black', font=('Arial',12))
label7.grid(row=8,column=0,padx=5,pady=10, sticky=E)
entry7 = Entry(window, width = 10, font=('Arial',12))
entry7.grid(row=8, column=1, sticky=W)

entry5.configure(state="disabled")
entry5.update()
entry6.configure(state="disabled")
entry6.update()    
entry7.configure(state="disabled")
entry7.update()    

# Additional output information
label2 = Label(window, text='Additional output: ', fg='black', font=('Arial',12))
label2.grid(row=9,column=0,padx=5,pady=10, sticky=W)

additional_output1_var = StringVar()
checkbox11 = Checkbutton(window, text = "SOFC AC power output",  variable = additional_output1_var, onvalue = "SOFC AC power", fg='black', font=('Arial',12))
checkbox11.deselect()
checkbox11.grid(row=9,column=1,padx=5,pady=10, sticky=W)

additional_output2_var = StringVar()
checkbox21 = Checkbutton(window, text = "SOFC oxygen utilisation",  variable = additional_output2_var, onvalue = "SOFC oxygen utilisation", fg='black', font=('Arial',12))
checkbox21.deselect()
checkbox21.grid(row=10,column=1,padx=5,pady=10, sticky=W)

additional_output3_var = StringVar()
checkbox31 = Checkbutton(window, text = "SOFC current density",  variable = additional_output3_var, onvalue = "SOFC current density", fg='black', font=('Arial',12))
checkbox31.deselect()
checkbox31.grid(row=11,column=1,padx=5,pady=10, sticky=W)

additional_output4_var = StringVar()
checkbox21 = Checkbutton(window, text = "Air compressor power",  variable = additional_output4_var, onvalue = "Air compressor power", fg='black', font=('Arial',12))
checkbox21.deselect()
checkbox21.grid(row=9,column=2,padx=5,pady=10, sticky=W)

additional_output5_var = StringVar()
checkbox22 = Checkbutton(window, text = "Transmitted heat air heater",  variable = additional_output5_var, onvalue = "Transmitted heat air heater", fg='black', font=('Arial',12))
checkbox22.deselect()
checkbox22.grid(row=10,column=2,padx=5,pady=10, sticky=W)

additional_output6_var = StringVar()
checkbox23 = Checkbutton(window, text = "Heat regeneration sat ST",  variable = additional_output6_var, onvalue = "Heat regeneration sat ST", fg='black', font=('Arial',12))
checkbox23.deselect()
checkbox23.grid(row=11,column=2,padx=5,pady=10, sticky=W)

additional_output7_var = StringVar()
checkbox24 = Checkbutton(window, text = "Heat regeneration HW",  variable = additional_output7_var, onvalue = "Heat regeneration HW", fg='black', font=('Arial',12))
checkbox24.deselect()
checkbox24.grid(row=12,column=2,padx=5,pady=10, sticky=W)

# Buttons
start_button = Button(window,text="Start", command = button_command_start, width = 20)
start_button.grid(row=13,column=0)

clear_button = Button(window,text="Clear", command = button_command_clear, width = 20)
clear_button.grid(row=13,column=1,sticky=W)



window.mainloop()